import openpyxl

def inspect_file(filepath, out):
    out.write(f"\n=========================================\nInspecting: {filepath}\n=========================================\n")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        out.write(f"Sheet names: {wb.sheetnames}\n")
        for name in wb.sheetnames:
            ws = wb[name]
            out.write(f"\nSheet: {name} (Max row: {ws.max_row}, Max col: {ws.max_column})\n")
            # Print headers and first 4 rows
            for r in range(1, min(ws.max_row + 1, 6)):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                out.write(f"Row {r}: {row_vals}\n")
    except Exception as e:
        out.write(f"Error inspecting {filepath}: {str(e)}\n")

with open("inspect_out.txt", "w", encoding="utf-8") as out:
    inspect_file("automated_test/DAST_Report_2026-06-10T07-12-39.xlsx", out)
